from shutil import copy
from openpyxl import Workbook, load_workbook
from os.path import exists
from os import mkdir
from pprint import pprint as pp
import datetime


def timestamp(func):
    """
    Декоратор timestamp
    :param func: функция декоратора
    :return: возвращает время затраченное на обработку функции
    """

    def wrapper(*args):
        start = datetime.datetime.now()
        func(*args)
        pp(f"Затрачено времени на генерацию: {datetime.datetime.now() - start}")

    return wrapper


class DataBase:
    def __init__(self, path):
        """
        Создаем базу данных, вызываем загрузку в конструкторе
        :param path: Путь до списка сотрудников в формате xlsx
        """
        self.path = path
        self._db = []
        self._load_db()

    def _load_db(self):
        """
        Загрузка БД xlsx с Лист1
        :return:
        """
        wb = load_workbook(self.path)
        sheet_ranges = wb["Лист1"]
        for row in sheet_ranges.iter_rows():
            for cell in row:
                self._db.append(cell.value)

    def _dir_create(self, path):
        """
        Проверка наличия папки, если она не создана - создаем
        :param path: путь к папке
        """
        if not exists(path):
            mkdir(path)

    def __repr__(self):
        return "\n".join(self._db)

    @property
    def db(self):
        """
        Геттер
        :return: защищенная переменная _db
        """
        return self._db

    @timestamp
    def create_files(self, date, path_to_copy, path_to_template):
        self._dir_create(path_to_template)
        for item in self._db:
            try:
                copy(path_to_copy, fr"{path_to_template}\{date}_{item}.xlsx")
                pp(f"Создание файла: {date}_{item} - создан!")
            except (FileNotFoundError, InterruptedError):
                pp(f"Ошибка при создании файла: {date}_{item}")
