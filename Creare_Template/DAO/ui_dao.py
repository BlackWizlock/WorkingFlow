from tkinter import *
from tkinter import filedialog
from tkinter import ttk

import json

from shutil import copy

from openpyxl import Workbook, load_workbook

from os.path import exists
from os import mkdir
import datetime

from pprint import pprint as pp


def timestamp(func):
    """
    Декоратор timestamp
    :param func: функция декоратора
    :return: возвращает время затраченное на обработку функции
    """

    def wrapper(*args):
        start = datetime.datetime.now()
        func(*args)
        pp(f"Затрачено времени на генерацию: {datetime.datetime.now() - start}")

    return wrapper


class Interface:
    def __init__(self):
        self.config = self._config_loader()

    @staticmethod
    def _config_loader():
        with open("default_values.json", "r", encoding="UTF-8") as f:
            return json.load(f)

    def _config_writer(self, setting, input_from_user):
        self.config[setting] = input_from_user
        with open("default_values.json", "w", encoding="UTF-8") as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)


class DataBase(Interface):
    def __init__(self):
        super().__init__()
        """
        Создаем базу данных, вызываем загрузку в конструкторе
        :param path: Путь до списка сотрудников в формате xlsx
        """
        self._db = []
        self._load_db()

    def _load_db(self):
        """
        Загрузка БД xlsx с Лист1
        :return:
        """
        wb = load_workbook(self.config["path_1"])
        sheet_ranges = wb["Лист1"]
        for row in sheet_ranges.iter_rows():
            for cell in row:
                self._db.append(cell.value)

    def _dir_create(self):
        """
        Проверка наличия папки, если она не создана - создаем
        :param path: путь к папке
        """
        if not exists(self.config["path_2"]):
            mkdir(self.config["path_2"])

    def __repr__(self):
        return "\n".join(self._db)

    @property
    def db(self):
        """
        Геттер
        :return: защищенная переменная _db
        """
        return self._db

    @timestamp
    def create_files(self, date, path_to_copy, path_to_template):
        self._dir_create()
        for item in self._db:
            try:
                path_to_templates += fr"\{date}_Табели"
                path_to_file_to_be_copy += fr"\_Шаблон_Табеля_{date}.xlsx"
                copy(path_to_copy, fr"{path_to_template}\{self.config['combo_4']}_{item}.xlsx")
                pp(f"Создание файла: {self.config['combo_4']}_{item} - создан!")
            except (FileNotFoundError, InterruptedError):
                pp(f"Ошибка при создании файла: {date}_{item}")


class Window(DataBase):
    def __init__(self, width, height, title="Создание табелей", resizable=(False, False), icon=None):
        super().__init__()
        self.root = Tk()
        self.root.title(title)
        self.root.geometry(f"{width}x{height}+200+200")
        self.root.resizable(resizable[0], resizable[1])
        if icon:
            self.root.iconbitmap(icon)
        self.first_line = Frame(self.root)
        self.second_line = Frame(self.root)
        self.third_line = Frame(self.root)
        self.btn_line = Frame(self.root)
        self.label_1 = Label(self.first_line, text="Список сотрудников:", width=20)
        self.path_1 = Entry(self.first_line, width=40)
        self.btn_1 = Button(self.first_line, text="Открыть", width=10, command=self._btn_1)
        self.label_2 = Label(self.second_line, text=fr'Папка "Табели":', width=20)
        self.path_2 = Entry(self.second_line, width=40)
        self.btn_2 = Button(self.second_line, text="Открыть", width=10, command=self._btn_2)
        self.label_3 = Label(self.third_line, text="Шаблон:", width=20)
        self.path_3 = Entry(self.third_line, width=40)
        self.btn_3 = Button(self.third_line, text="Открыть", width=10, command=self._btn_3)
        self.label_4 = Label(self.btn_line, text="На какой месяц:", width=20)
        self.combo_4 = ttk.Combobox(self.btn_line, values=[
                "2201", "2202", "2203", "2204", "2205", "2206", "2207", "2208", "2209", "2210", "2211", "2212"
        ])
        self.btn_4 = Button(self.btn_line, text="Создать табели", width=30, command=self._btn_4)
        self.path_to_file_1 = self.config["path_1"]
        self.path_to_file_2 = self.config["path_2"]
        self.path_to_file_3 = self.config["path_3"]

    def draw_widgets(self):
        self.first_line.pack()
        self.second_line.pack()
        self.third_line.pack()
        self.btn_line.pack()
        self.label_1.pack(side=LEFT, padx=5, pady=5)
        self.path_1.pack(side=LEFT, padx=5, pady=5)
        self.path_1.insert(0, self._default_values("path_1"))
        self.btn_1.pack(side=LEFT, padx=5, pady=5)
        self.label_2.pack(side=LEFT, padx=5, pady=5)
        self.path_2.pack(side=LEFT, padx=5, pady=5)
        self.path_2.insert(0, self._default_values("path_2"))
        self.btn_2.pack(side=LEFT, padx=5, pady=5)
        self.label_3.pack(side=LEFT, padx=5, pady=5)
        self.path_3.pack(side=LEFT, padx=5, pady=5)
        self.path_3.insert(0, self._default_values("path_3"))
        self.btn_3.pack(side=LEFT, padx=5, pady=5)
        self.label_4.pack(side=LEFT, padx=5, pady=5)
        self.combo_4.pack(side=LEFT, padx=5, pady=5)
        self.combo_4.insert(0, self._default_values("combo_4"))
        self.btn_4.pack(side=LEFT, padx=5, pady=5)

    def run(self):
        self.draw_widgets()
        self.root.mainloop()

    def _default_values(self, setting):
        for key, value in self.config.items():
            if setting == key and value:
                return value
        return ""

    def _btn_1(self):
        self.path_1.delete(0, END)
        file_1 = filedialog.askopenfilename(filetypes=(("XLSX", "*.xlsx"), ("XLS", "*.xls")))
        self.path_1.insert(1, str(file_1))
        self.path_to_file_1 = file_1

    def _btn_2(self):
        self.path_2.delete(0, END)
        file_2 = filedialog.askdirectory()
        self.path_2.insert(1, str(file_2))
        self.path_to_file_2 = file_2

    def _btn_3(self):
        self.path_3.delete(0, END)
        file_3 = filedialog.askdirectory()
        self.path_3.insert(1, str(file_3))
        self.path_to_file_3 = file_3

    def _btn_4(self):
        if self.combo_4.get() and self.path_to_file_1 and self.path_to_file_2 and self.path_to_file_3:
            self._config_writer("path_1", self.path_to_file_1)
            self._config_writer("path_2", self.path_to_file_2)
            self._config_writer("path_3", self.path_to_file_3)
            self._config_writer("combo_4", self.combo_4.get())
            # Window.create_files()