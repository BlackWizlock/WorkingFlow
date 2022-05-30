from openpyxl import Workbook, load_workbook
from pprint import pprint as pp
from os.path import exists
from os import mkdir
from shutil import copy
import datetime


# создаем базу данных
def dir_create(path):
    if not exists(path):
        mkdir(path)


def timestamp(func):
    def wrapper(*args):
        start = datetime.datetime.now()
        func(*args)
        pp(f"Затрачено времени на генерацию: {datetime.datetime.now() - start}")

    return wrapper


class DataBase:
    def __init__(self, path):
        """
        Создаем базу данных, вызываем загрузку в конструкторе
        :param path: Путь до списка сотрудников в формате xlsx
        """
        self.path = path
        self._db = []
        self._load_db()

    def _load_db(self):
        """
        Загрузка БД xlsx с Лист1
        :return:
        """
        wb = load_workbook(self.path)
        sheet_ranges = wb["Лист1"]
        for row in sheet_ranges.iter_rows():
            for cell in row:
                self._db.append(cell.value)

    def __repr__(self):
        return "\n".join(self._db)

    @property
    def db(self):
        """
        Геттер
        :return: защищенная переменная _db
        """
        return self._db

    @timestamp
    def create_files(self, date, path_to_copy, path_to_template):
        for item in self._db:
            try:
                copy(path_to_copy, fr"{path_to_template}\{date}_{item}.xlsx")
                pp(f"Создание файла: {date}_{item} - создан!")
            except:
                pp(f"Ошибка при создании файла: {date}_{item}")


def main():
    while True:
        date = input('Введите дату для создания шаблона (ГГММ):')
        if date.isdigit():
            break
        else:
            pp("Ошибка в формате ввода (ГГММ), повторите попытку")

    path_to_workers = r"\\pnk2.local\resources\Проектный отдел\Общие\Администрирование\Список сотрудников ПД\Список сотрудников ПД.xlsx"
    path_to_templates = fr"\\pnk2.local\resources\Проектный отдел\Общие\Администрирование\Табель\{date}_Табели"
    path_to_file_to_be_copy = fr"\\pnk2.local\resources\Проектный отдел\Общие\Администрирование\Табель\Шаблоны табелей\_Шаблон_Табеля_{date}.xlsx"

    database = DataBase(path_to_workers)  # создаем папку - если её нет
    dir_create(path_to_templates)  # копируем шаблон и переименовываем под БД

    database.create_files(date, path_to_file_to_be_copy, path_to_templates)


if __name__ == '__main__':
    main()
