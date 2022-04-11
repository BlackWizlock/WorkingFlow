import shutil, datetime, time
from openpyxl import load_workbook, Workbook


def get_next_tmpl_date() -> str:
	ts = time.time()
	usr_mnt = int(datetime.datetime.fromtimestamp(ts).strftime('%m')) + 1
	usr_mnt = str(usr_mnt).zfill(2)
	st = datetime.datetime.fromtimestamp(ts).strftime('%y')
	return st + usr_mnt


def get_worker_name(path: str) -> list:
	wb = load_workbook(path)  # Work Book
	ws = wb['Лист1']  # Work Sheet
	column = ws['A']  # Column
	column_list = [column[x].value for x in range(len(column))]
	return column_list


def make_final_fn(name_fl: str) -> str:
	return f'\\\\pnk2.local\\resources\\Проектный отдел\\Общие\\Администрирование\\Табель\\' \
		   f'{get_next_tmpl_date()}_Табели\\{get_next_tmpl_date()}_{name_fl}.xlsx'


def generate_cal():
	return ["Проект:"] + [f'{str(x).zfill(2)}.04.2022' for x in range(1, 31)]



def main():
	file_to_be_copy = f'_Шаблон_Табеля_{get_next_tmpl_date()}.xlsx'
	tmpl_dir = f'\\\\pnk2.local\\resources\\Проектный отдел\\Общие\\Администрирование\\Табель\\Шаблоны ' \
			   f'табелей\\{file_to_be_copy}'
	worker_dir_fln = f'\\\\pnk2.local\\resources\\Проектный отдел\\Общие\\Администрирование\\Список сотрудников ' \
					 f'ПД\\Список сотрудников ПД.xlsx'

	for name in get_worker_name(worker_dir_fln):
		print(f'{name} was added')
		shutil.copyfile(tmpl_dir, make_final_fn(name))

	print('Finished process')


if __name__ == '__main__':
	main()
