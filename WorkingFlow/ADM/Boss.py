import datetime
import json
import os
import pathlib
import re
import sys
import time
from queue import Queue

import openpyxl
import openpyxl.styles
from PyQt5.QtGui import QFont
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter

import Main
import about
import default_window
import traceback
import cProfile

from PyQt5.QtCore import (QDir, QDate, pyqtSignal, QThread)
from PyQt5.QtWidgets import (QMainWindow, QApplication, QFileDialog, QMessageBox, QDialog, QButtonGroup, QLineEdit,
                             QSizePolicy, QPushButton, QLabel, QVBoxLayout, QTextEdit, QScrollArea, QWidget)


class ScrollMessageBox(QMessageBox):
    def __init__(self, text, *args, **kwargs):
        QMessageBox.__init__(self, *args, **kwargs)
        text_edit = QTextEdit(self)
        text_edit.setPlainText(text)
        lay = QVBoxLayout()
        lay.addWidget(text_edit)
        self.layout().addWidget(text_edit, 0, 0, 1, self.layout().columnCount())
        self.setStyleSheet("QTextEdit{min-width:460 px; min-height: 600px}")
        self.setWindowTitle('Аномалии')


class DefaultWindowWindow(QDialog, default_window.Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        try:
            with open(pathlib.Path(pathlib.Path.cwd(), 'Настройки.txt'), "r", encoding='utf-8-sig') as f:
                self.data = json.load(f)
        except FileNotFoundError:
            pass
        self.name_eng = ['path_boss', 'path_worker', 'path_finish_file', 'path_worker_list', 'password']
        self.name_rus = ['Путь к файлу с проектами', 'Путь к файлам работников', 'Путь для конечного файла',
                         'Путь к списку работников', 'Пароль']
        self.buttongroup_add = QButtonGroup()
        self.buttongroup_add.buttonClicked[int].connect(self.add_button_clicked)
        self.pushButton_ok.clicked.connect(self.accept)
        self.pushButton_cancel.clicked.connect(self.cancel)
        self.i = 0
        self.line = {}
        self.name = {}
        self.button = {}
        for el in self.name_rus:
            self.line[self.i] = QLabel(self.frame_sett)
            self.line[self.i].setText(el)
            self.line[self.i].setFont(QFont("Times", 12, QFont.Light))
            self.line[self.i].setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            self.line[self.i].setDisabled(True)
            self.gridLayout_8.addWidget(self.line[self.i], self.i, 0)
            self.name[self.i] = QLineEdit(self.frame_sett)
            try:
                self.name[self.i].setText(self.data[self.name_eng[self.name_rus.index(el)]])
            except KeyError:
                pass
            self.name[self.i].setFont(QFont("Times", 12, QFont.Light))
            self.name[self.i].setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            self.name[self.i].setDisabled(True)
            self.gridLayout_8.addWidget(self.name[self.i], self.i, 1)
            self.button[self.i] = QPushButton("Изменить", self.frame_sett)  # Создаем кнопку
            self.button[self.i].setFont(QFont("Times", 12, QFont.Light))  # Размер шрифта
            self.button[self.i].setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)  # Размеры виджета
            self.buttongroup_add.addButton(self.button[self.i], self.i)  # Добавляем в группу
            self.gridLayout_8.addWidget(self.button[self.i], self.i, 2)  # Добавляем в фрейм по месту
            self.i += 1  # Увеличиваем счетчик

    def add_button_clicked(self, number):  # Если кликнули по кнопке
        self.name[number].setEnabled(True)

    def cancel(self):  # Если кликнули по кнопке cancel
        self.close()

    def accept(self):
        for el in self.name:
            if self.name[el].isEnabled():
                if self.name[el].text():
                    self.data[self.name_eng[self.name_rus.index(self.line[el].text())]] = self.name[el].text()
                else:
                    self.data.pop(self.name_eng[self.name_rus.index(self.line[el].text())], None)
        with open(pathlib.Path(pathlib.Path.cwd(), 'Настройки.txt'), 'w', encoding='utf-8-sig') as f:
            json.dump(self.data, f, ensure_ascii=False, sort_keys=True, indent=4)
        self.close()


class AboutWindow(QDialog, about.Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


def work_list(path_worker_list):
    wb = openpyxl.load_workbook(path_worker_list)  # Откроем книгу.
    ws = wb.active  # Делаем активным первый лист.
    i = 1
    workers = []
    while True:
        if ws.cell(i, 1).value:
            workers.append(ws.cell(i, 1).value)
        else:
            break
        i += 1
    wb.close()
    return workers


def chek_file(element, date):
    errors = []
    if '~' not in element:
        wb = openpyxl.load_workbook(element)
        ws = wb.active
        for el in range(2, ws.max_column):
            try:
                time_year = str(ws.cell(1, el).value.year)[2:]
                time_month = ws.cell(1, el).value.month
                t = time_year + '0' + str(time_month) if time_month < 10 else \
                    time_year + str(time_month)
                if t != date:
                    errors += ['Неверно указано время в файле работника ' + element]
                    break
            except AttributeError:
                break
    return errors


class WorkerAnalytics(QThread):
    status = pyqtSignal(str)
    messageChanged = pyqtSignal(str, str)

    def __init__(self, path_worker, path_worker_list, date_analytics, path_finish_file, finish_date):
        QThread.__init__(self)
        self.path_worker = path_worker
        self.date_analytics = date_analytics
        self.path_worker_list = path_worker_list
        self.path_finish_file = path_finish_file
        self.finish_date = finish_date

    def run(self):
        self.status.emit('Считываем список работников')
        wb = openpyxl.load_workbook(self.path_worker_list)  # Откроем книгу.
        ws = wb.active  # Делаем активным первый лист.
        i = 1
        workers = []
        while True:
            if ws.cell(i, 1).value:
                workers.append(ws.cell(i, 1).value)
            else:
                break
            i += 1
        wb.close()
        os.chdir(self.path_worker)
        # for el in os.listdir():
        #     if self.date_analytics == el.partition('_')[0]:
        #         os.chdir(self.path_worker + '/' + el)
        #         break
        worker_time = {}
        all_time = 0
        days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        day = days[int(self.date_analytics[2:]) - 1]
        flag_for_time = True

        os.chdir(self.path_worker)
        file_worker = [i for i in os.listdir()]
        workers = work_list(self.path_worker_list)
        errors = []
        pass_file = []
        date = self.date_analytics
        for folder in file_worker:
            err = None
            if folder.partition('_')[0] == date:
                if (self.path_worker + "\\" + folder).endswith(".lnk"):
                    pass
                if os.path.isdir(self.path_worker + "\\" + folder):
                    os.chdir(self.path_worker + '\\' + folder)
                    file_exel = [i.partition('_')[2][:-5] for i in os.listdir() if i.endswith('.xlsm') or i.endswith('.xlsx')]
                    err = ['Отсутствует табель ' + date + '_' + i for i in workers if i not in file_exel]
                    if err:
                        errors += err
                    err = ['Табель ' + date + '_' + i + ' отсутствует в списке работников' for i in file_exel
                           if i not in workers]
                    if err:
                        errors += err
                    for element in [i for i in os.listdir() if i.endswith('.xlsm') or i.endswith('.xlsx')]:
                        self.status.emit('Проверка заполнения файлов ' + element)
                        err = chek_file(element, date)
                        if err:
                            errors += err
                            pass_file.append(element)
        if errors:
            self.messageChanged.emit('Скролл', '\n'.join(errors))
        for el in os.listdir():
            if (el.endswith('.xlsm') or el.endswith('.xlsx')) and ('~' not in el) and (el not in pass_file):
                self.status.emit('Считываем время работника: ' + el.partition('_')[2].rpartition('.')[0])
                sum_hour = 0
                wb = openpyxl.load_workbook(el, data_only=True)  # Откроем книгу.
                ws = wb.active
                for i in range(2, day + 2):
                    try:
                        sum_hour += ws.cell(2, i).value
                    except TypeError:
                        pass
                    if flag_for_time:
                        try:
                            if ws.cell(1, i).value.weekday() not in [5, 6]:
                                all_time += 8
                        except AttributeError:
                            date_ = ws.cell(1, i).value
                            year_ = int(date_.rpartition('.')[2])
                            month_ = int(date_.rpartition('.')[0].partition('.')[2])
                            day_ = int(date_.partition('.')[0])
                            date_ = datetime.date(year_, month_, day_)
                            if date_ not in [5, 6]:
                                all_time += 8
                worker_time[el.rpartition('.')[0].partition('_')[2]] = sum_hour
            wb.close()
            flag_for_time = False
        self.status.emit('Заполняем файл руководителя')
        wb = openpyxl.Workbook()  # Откроем книгу.
        ws = wb.active  # Делаем активным первый лист.
        ws['A1'] = 'Работник'
        ws['B1'] = 'Время'
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws['B1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        thin = openpyxl.styles.Side(border_style="thin", color="000000")
        ws['A1'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
        ws['B1'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
        i = 2
        for el in workers:
            self.status.emit('Заполняем файл руководителя. Работник: ' + el)
            ws.cell(i, 1).value = el
            ws.cell(i, 1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            try:
                ws.cell(i, 2).value = worker_time[el]
                ws.cell(i, 2).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                if worker_time[el] < all_time:
                    ws.cell(i, 1).fill = openpyxl.styles.PatternFill(start_color='ff0000', end_color='ff0000',
                                                                     fill_type="solid")
                    ws.cell(i, 2).fill = openpyxl.styles.PatternFill(start_color='ff0000', end_color='ff0000',
                                                                     fill_type="solid")
                    ws.cell(i, 1).font = openpyxl.styles.Font(bold=True, color='ffffff')
                    ws.cell(i, 2).font = openpyxl.styles.Font(bold=True, color='ffffff')
                    ws.cell(i, 1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    ws.cell(i, 2).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    ws.cell(i, 1).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(i, 2).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            except KeyError:
                flag = False
                for file_ in pass_file:
                    if re.findall(el, file_):
                        flag = True
                        break
                if flag:
                    ws.cell(i, 2).value = 'Неверно указано время'
                else:
                    ws.cell(i, 2).value = 'Нет табеля'
                ws.cell(i, 1).fill = openpyxl.styles.PatternFill(start_color='ff0000', end_color='ff0000',
                                                                 fill_type="solid")
                ws.cell(i, 2).fill = openpyxl.styles.PatternFill(start_color='ff0000', end_color='ff0000',
                                                                 fill_type="solid")
                ws.cell(i, 1).font = openpyxl.styles.Font(bold=True, color='ffffff')
                ws.cell(i, 2).font = openpyxl.styles.Font(bold=True, color='ffffff')
                ws.cell(i, 1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                ws.cell(i, 2).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                ws.cell(i, 1).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(i, 2).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            i += 1
        self.status.emit('Готово! Открытие файла аналитика руководителя по часам')
        os.chdir(self.path_finish_file)
        wb.save(filename='Аналитика руководителя по часам.xlsx')
        os.startfile('Аналитика руководителя по часам.xlsx')


class StartProc(QThread):
    status = pyqtSignal(str)
    messageChanged = pyqtSignal(str, str)

    def __init__(self, path_boss, path_worker, date_analytics, time_text, path_finish_file, start_year, start_month,
                 finish_year, finish_month, q, log, path_worker_list):
        QThread.__init__(self)
        self.path_boss = path_boss
        self.path_worker = path_worker
        self.date_analytics = date_analytics
        self.time_text = time_text
        self.path_finish_file = path_finish_file
        self.start_year = start_year
        self.start_month = start_month
        self.finish_year = finish_year
        self.finish_month = finish_month
        self.q = q
        self.log = log
        self.path_worker_list = path_worker_list

    def profile(func):
        """Decorator for run function profile"""

        def wrapper(*args, **kwargs):
            profile_filename = func.__name__ + '.prof'
            profiler = cProfile.Profile()
            result = profiler.runcall(func, *args, **kwargs)
            profiler.dump_stats(profile_filename)
            return result

        return wrapper

    def wait_for_dialog(self):
        while True:
            if self.ret == 1:
                pass
            elif self.ret == 2:
                break
            elif self.ret == 3:
                return

    @profile
    def run(self):

        def write(w, r, e, nw, elem, com):  # ws, row, el, name_worker, element, comments
            cell_column_def = 2
            for date_keys_def in nw[e][elem]:
                while True:
                    if date_keys_def == w.cell(1, cell_column_def).value:
                        w.cell(r, cell_column_def).value = nw[e][elem][date_keys_def]
                        if type(nw[e][elem][date_keys_def]) is float:
                            w.cell(r, cell_column_def).number_format = "0.0"
                            w.column_dimensions[get_column_letter(cell_column_def)].width = 3.5
                        break
                    cell_column_def += 1
            del nw[e][elem]
            cell_column_def = 2
            try:
                for date_keys_def in com[e][elem]:
                    while True:
                        if date_keys_def == w.cell(1, cell_column_def).value:
                            w.cell(r, cell_column_def).comment = com[e][elem][date_keys_def]
                            break
                        cell_column_def += 1
                del com[e][elem]
            except KeyError:
                pass

        try:
            self.status.emit('Считывание проектов')
            wb = openpyxl.load_workbook(self.path_boss, data_only=True)  # Откроем книгу.
            ws = wb.active  # Делаем активным первый лист.
            project_col = None
            j = 1
            while not project_col:
                if ws.cell(1, j).value == 'Шифр объекта':
                    project_col = j
                j += 1
            project = ['_ОТПУСК', '_ПРОСТОЙ', '_Управление (только для руководителя)', '_Отработка листов отклонений',
                       '_ИНОЕ (только по решению руководителя)']
            i = 2
            while project_col:
                if ws.cell(i, project_col).value:
                    project.append(ws.cell(i, project_col).value)
                else:
                    project_col = None
                i += 1
            wb.close()
            os.chdir(self.path_worker)
            if self.log:
                with open(self.path_worker + '\\log.txt', 'w') as f:
                    print('-----------------Проекты---------------------', file=f)
                    print(project, file=f)
            self.status.emit('Считывание времени работников')
            os.chdir(self.path_worker)
            file_worker = [i for i in os.listdir()]
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('-------------------Директории-------------------', file=f)
                    print(file_worker, file=f)
            name_worker = {}
            comments = {}
            if len(self.date_analytics) == 1:
                month_ap = '0' + str(self.date_analytics[0]) if self.date_analytics[0] < 10 \
                    else str(self.date_analytics[0])
                current_date = [self.finish_year[2:4] + month_ap]
            else:
                current_date = []
                year = int(self.start_year[2:4])
                month = int(self.start_month)
                self.date_analytics = []
                while True:
                    self.date_analytics.append(month)
                    month_ap = '0' + str(month) if month < 10 else str(month)
                    current_date.append(str(year) + month_ap)
                    if month == int(self.finish_month):
                        if year == int(self.finish_year[2:4]):
                            break
                    if month < 12:
                        month += 1
                    else:
                        month = 1
                        year += 1
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('-----------------Даты---------------------', file=f)
                    print(current_date, file=f)
                    print('-----------------Даты---------------------', file=f)
                    print(self.date_analytics, file=f)
            days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            day = [days[i - 1] for i in self.date_analytics]
            errors = []
            worker = {}
            worker_com = {}
            if time.localtime()[1] < 10:
                current_time = str(time.localtime()[0])[2:] + '0' + str(time.localtime()[1])
            else:
                current_time = str(time.localtime()[0])[2:] + str(time.localtime()[1])

            workers = work_list(self.path_worker_list)
            errors = []
            pass_file = []
            for date in current_date:
                for folder in file_worker:
                    err = None
                    if folder.partition('_')[0] == date:
                        if (self.path_worker + "\\" + folder).endswith(".lnk"):
                            pass
                        if os.path.isdir(self.path_worker + "\\" + folder):
                            os.chdir(self.path_worker + '\\' + folder)
                            file_exel = [i.partition('_')[2][:-5] for i in os.listdir() if (i.endswith('.xlsm')
                                         or i.endswith('.xlsx'))]
                            err = ['Отсутствует табель ' + date + '_' + i for i in workers if i not in file_exel]
                            if err:
                                errors += err
                            err = ['Табель ' + date + '_' + i + ' отсутствует в списке работников' for i in file_exel
                                   if i not in workers]
                            if err:
                                errors += err
                            for element in [i for i in os.listdir() if (i.endswith('.xlsm') or i.endswith('.xlsx'))]:
                                self.status.emit('Проверка заполнения файлов ' + element)
                                err = chek_file(element, date)
                                if err:
                                    errors += err
                                    pass_file.append(element)
            if errors:
                self.messageChanged.emit('Скролл', '\n'.join(errors))
            for date in current_date:
                for folder in file_worker:
                    ret = 0
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('-----------------Дата и папка---------------------', file=f)
                            print(date, '+', folder, file=f)
                    if folder.partition('_')[0] == date:
                        if (self.path_worker + "\\" + folder).endswith(".lnk"):
                            self.messageChanged.emit('Вопрос', folder +
                                                     ' является ярлыком и не будет учтён при обходе, продолжить?')
                            while True:
                                val = self.q.get()
                                if val == 1:
                                    pass
                                elif val == 2:
                                    ret = 1
                                    break
                                elif val == 3:
                                    self.status.emit('Прервано пользователем')
                                    return
                        if ret == 0:
                            if os.path.isdir(self.path_worker + "\\" + folder):
                                os.chdir(self.path_worker + '/' + folder)
                                file_exel = [i for i in os.listdir() if ((i.endswith('.xlsm')
                                                                          or i.endswith('.xlsx')) and ('~' not in i) and
                                                                         i not in pass_file)]
                                if self.log:
                                    with open(self.path_worker + '\\log.txt', 'a') as f:
                                        print('------------------Файлы работников--------------------', file=f)
                                        print(file_exel, file=f)
                                for file in file_exel:
                                    self.status.emit('Считывание времени работников ' + file.split('.')[0])
                                    try:
                                        if date == file.partition('_')[0]:
                                            wb = openpyxl.load_workbook(file, keep_links=True, data_only=True)
                                                                        # read_only=False, keep_vba=True, , keep_links=True, data_only=True)
                                            ws = wb.active  # Делаем активным первый лист.
                                            wname = file.partition('_')[2].partition('.')[0]
                                            if wname not in worker:
                                                worker[wname] = {}
                                                worker_com[wname] = {}
                                            worker_project = {}
                                            comment = {}
                                            max_row = ws.max_row
                                            for i in range(3, max_row + 1):
                                                flag = 0
                                                for j in range(2, day[current_date.index(date)] + 2):
                                                    if ws.cell(i, j).value:
                                                        if re.findall(r'ref', ws.cell(i, 1).value.lower()):
                                                            proj = project[i - 1]
                                                        else:
                                                            proj = ws.cell(i, 1).value
                                                        if ws.cell(i, 1).value not in worker[wname]:
                                                            worker[wname][proj] = {}
                                                            worker_com[wname][proj] = {}
                                                        if flag == 0:
                                                            worker_project[proj] = {}
                                                            comment[proj] = {}
                                                            flag = 1
                                                        try:
                                                            if '\\n' in ws.cell(1, j).value:
                                                                date_ = ws.cell(1, j).value.partition('\\n')[0]
                                                                year_ = int(date_.rpartition('.')[2])
                                                                month_ = int(date_.rpartition('.')[0].partition('.')[2])
                                                                day_ = int(date_.partition('.')[0])
                                                                date_ = datetime.date(year_, month_, day_)
                                                            else:
                                                                date_ = ws.cell(1, j).value
                                                                year_ = int(date_.rpartition('.')[2])
                                                                month_ = int(date_.rpartition('.')[0].partition('.')[2])
                                                                day_ = int(date_.partition('.')[0])
                                                                date_ = datetime.date(year_, month_, day_)
                                                        except TypeError:
                                                            date_ = ws.cell(1, j).value
                                                            date_ = datetime.date(date_.year, date_.month, date_.day)
                                                        if date_.month < 10:
                                                            date_now = str(date_.year)[2:] + '0' + str(date_.month)
                                                        else:
                                                            date_now = str(date_.year)[2:] + str(date_.month)
                                                        if date != date_now:
                                                            self.status.emit(
                                                                'Ошибка в дате в файле работника ' + file.split('.')[0])
                                                            return
                                                        if type(ws.cell(i, j).value) is str:
                                                            try:
                                                                value = round(float(ws.cell(i, j).value), 1)
                                                            except ValueError:
                                                                value = ws.cell(i, j).value
                                                        else:
                                                            value = ws.cell(i, j).value
                                                        worker_project[proj][date_] = value
                                                        worker[wname][proj][date_] = value
                                                        if ws.cell(i, j).comment:
                                                            worker_com[wname][proj][date_] \
                                                                = ws.cell(i, j).comment
                                                            comment[proj][date_] = \
                                                                ws.cell(i, j).comment
                                            wb.close()
                                            # if file.partition('_')[0] < current_time:
                                            #     wb = openpyxl.open(filename=file, data_only=False, keep_links=True,
                                            #                    keep_vba=True)
                                            #     spisok_ws = wb.sheetnames
                                            #     ws = wb[spisok_ws[0]]  # Делаем активным первый лист.
                                            #     ws.protection.set_password('123')
                                            # wb.save(filename=file)
                                            # wb.close()
                                            name_worker[file.split('.')[0]] = worker_project
                                            comments[file.split('.')[0]] = comment
                                    except Exception:
                                        errors.append('Ошибка в файле работника ' + file)
                                        print(traceback.format_exc())
                                        if self.log:
                                            with open(self.path_worker + '\\log.txt', 'a') as f:
                                                print('--------------------------------------', file=f)
                                                print('Ошибка в файле работника ' + file)
                                                print('--------------------------------------', file=f)
                                                print(traceback.format_exc(), file=f)
            os.chdir(self.path_worker)
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('--------------------Проекты------------------', file=f)
                    print(name_worker, file=f)
                    print('--------------------Комментарии------------------', file=f)
                    print(comments, file=f)
            self.status.emit('Формирование файла аналитика руководителя')
            name_file = str(datetime.date.today()) + '_Аналитика руководителя.xlsx'
            wb = openpyxl.Workbook()  # Откроем книгу.
            thin = openpyxl.styles.Side(border_style="thin", color="000000")
            ws = wb.active  # Делаем активным первый лист.
            ws.title = 'По работникам (общее)'
            col = 2
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 15
            ws['A1'].value = 'Работник (проекты)'
            ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            ws['A1'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['A1'].font = openpyxl.styles.Font(name='Calibri', size=14, bold=True)
            ws['B1'].value = 'Сумма часов'
            ws['B1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            ws['B1'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            ws['B1'].font = openpyxl.styles.Font(name='Calibri', size=14, bold=True)
            for element in name_worker:
                self.status.emit('Формирование файла аналитика руководителя (по часам): ' + element)
                if self.log:
                    with open(self.path_worker + '\\log.txt', 'a') as f:
                        print('------------------Работник общее--------------------', file=f)
                        print(element, file=f)
                ws.merge_cells(start_row=col, start_column=1, end_row=col, end_column=2)
                ws.cell(col, 1).value = element
                ws.cell(col, 1).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(col, 2).border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(col, 1).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                col += 1
                for proj in name_worker[element]:
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('-------------------Проекты общее-------------------', file=f)
                            print(proj, file=f)
                    ws.cell(col, 1).value = proj
                    ws.cell(col, 1).border = openpyxl.styles.Border(right=thin)
                    ws.cell(col, 1).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
                    try:
                        ws.cell(col, 2).value = sum(name_worker[element][proj].values())
                    except TypeError:
                        sum_ = 0
                        for val_ in name_worker[element][proj]:
                            if type(name_worker[element][proj][val_]) is not str:
                                sum_ += name_worker[element][proj][val_]
                        ws.cell(col, 2).value = sum_
                    ws.cell(col, 2).border = openpyxl.styles.Border(right=thin)
                    ws.cell(col, 2).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    col += 1
            ws.cell(col, 1).border = ws.cell(col, 2).border = openpyxl.styles.Border(top=thin)
            wb.save(filename=name_file)
            wb.close()
            wb = openpyxl.open(filename=name_file)
            wb.create_sheet('По работникам (часы)')
            ws = wb['По работникам (часы)']
            ws['A1'] = 'Работник:'
            ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            ws.column_dimensions['A'].width = 40
            ws['A2'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            col = 2
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('-----------------Работник (по часам)---------------------', file=f)
            for element in day:
                for date_for_month in range(1, element + 1):  # Изменяем оформление первой строки.
                    cell_header = ws.cell(1, col)
                    ws.column_dimensions[get_column_letter(col)].width = 3
                    cell_header.alignment = openpyxl.styles.Alignment(textRotation=90)
                    cell_header.border = openpyxl.styles.Border(bottom=thin)
                    cell_header.value = datetime.date(int(self.time_text[0]), self.date_analytics[day.index(element)],
                                                      date_for_month)
                    ws.cell(2, col).border = openpyxl.styles.Border(bottom=thin)
                    col += 1
            row = 3
            for element in worker:
                try:
                    self.status.emit('Формирование файла аналитика руководителя (по работникам): ' + element)
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('------------------Работник по часам--------------------', file=f)
                            print(element, file=f)
                    ws.cell(row, 1).value = element
                    ws.cell(row, 1).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
                    ws.cell(row, 1).border = openpyxl.styles.Border(right=thin, top=thin)
                    row += 1
                    for proj in worker[element]:
                        if self.log:
                            with open(self.path_worker + '\\log.txt', 'a') as f:
                                print('-------------------Проекты по часам-------------------', file=f)
                                print(proj, file=f)
                        ws.cell(row, 1).value = proj
                        ws.cell(row, 1).alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
                        ws.cell(row, 1).border = openpyxl.styles.Border(right=thin)
                        for el in worker[element][proj]:
                            for i in range(2, ws.max_column):
                                if ws.cell(1, i).value == el:
                                    ws.cell(row, i).value = worker[element][proj][el]
                                    break
                        if worker_com[element][proj]:
                            for el in worker_com[element][proj]:
                                for i in range(2, ws.max_row):
                                    if ws.cell(1, i).value == el:
                                        ws.cell(row, i).comment = worker_com[element][proj][el]
                                        break
                        row += 1
                except Exception:
                    errors.append('Ошибка при заполнении файла аналитики у работника ' + element + ' проект ' + proj)
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('--------------------------------------', file=f)
                            print('Ошибка в файле работника ' + element)
                            print('--------------------------------------', file=f)
                            print(traceback.format_exc(), file=f)
            max_row = ws.max_row
            max_col = ws.max_column
            col = 2
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('--------------------------------------', file=f)
                    print('Суммирование', file=f)
            while col <= max_col:
                ws.cell(2, col).value = '= SUM(' + get_column_letter(col) + str(3) + ':' \
                                        + get_column_letter(col) + str(max_row) + ')'
                if ws.cell(1, col).value.weekday() in [5, 6]:
                    for cell_row in range(1, max_row + 1):
                        ws.cell(cell_row, col).fill = openpyxl.styles.PatternFill(start_color='dedede',
                                                                                  end_color='dedede',
                                                                                  fill_type="solid")
                else:
                    ws.cell(2, col).fill = openpyxl.styles.PatternFill(start_color='90e896', end_color='90e896',
                                                                       fill_type="solid")
                col += 1
            wb.save(filename=name_file)
            wb.close()
            wb = openpyxl.open(filename=name_file)
            wb.create_sheet('По проектам')
            ws = wb['По проектам']
            ws['A1'] = 'Проект:'
            ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            ws.column_dimensions['A'].width = 40
            ws['A2'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            col = 2
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('-----------------Проекты---------------------', file=f)
            for element in day:
                for date_for_month in range(1, element + 1):  # Изменяем оформление первой строки.
                    cell_header = ws.cell(1, col)
                    ws.column_dimensions[get_column_letter(col)].width = 3
                    cell_header.alignment = openpyxl.styles.Alignment(textRotation=90)
                    cell_header.border = openpyxl.styles.Border(bottom=thin)
                    cell_header.value = datetime.date(int(self.time_text[0]), self.date_analytics[day.index(element)],
                                                      date_for_month)
                    ws.cell(2, col).border = openpyxl.styles.Border(bottom=thin)
                    col += 1
            row = 3
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('-------------------Дни-------------------', file=f)
            for element in project:
                self.status.emit('Формирование файла аналитика руководителя (по проектам): ' + element)
                try:
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('------------------Проекты для файла--------------------', file=f)
                            print(element, file=f)
                    if row <= ws.max_row:
                        row = ws.max_row + 1
                    if not ws.cell(row - 1, 1).value and row != 3:
                        row -= 1
                    cell_header = ws.cell(row, 1)
                    cell_start = row + 1
                    cell_header.border = openpyxl.styles.Border(right=thin, top=thin)
                    cell_header.value = element
                    for el in name_worker:
                        if self.log:
                            with open(self.path_worker + '\\log.txt', 'a') as f:
                                print('------------------Работник--------------------', file=f)
                                print(el, file=f)
                        if element in name_worker[el].keys():
                            row += 1
                            len_cell = cell_start
                            while True:
                                try:
                                    if '/' in ws.cell(len_cell, 1).value:
                                        break
                                    else:
                                        if el.partition('_')[2] == ws.cell(len_cell, 1).value:
                                            row = len_cell
                                            break
                                        else:
                                            len_cell += 1
                                except TypeError:
                                    break
                            ws.cell(row, 1).value = el.partition('_')[2]
                            ws.cell(row, 1).border = openpyxl.styles.Border(right=thin)
                            ws.cell(row, 1).alignment = openpyxl.styles.Alignment(horizontal="right")
                            write(ws, row, el, name_worker, element, comments)
                    row += 1
                except Exception:
                    errors.append('Ошибка при заполнении файла аналитики в проекте ' + element + ' работник ' + el)
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('--------------------------------------', file=f)
                            print('Ошибка в файле работника ' + file)
                            print('--------------------------------------', file=f)
                            print(traceback.format_exc(), file=f)
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('--------------------------------------', file=f)
                    print(worker_project, file=f)
                    print('--------------------------------------', file=f)
                    print(comments, file=f)
            cell_start = row
            flag = 0
            # wb.save(filename=name_file)
            # wb.close()
            # wb = openpyxl.open(filename=name_file)
            ws = wb['По проектам']
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('--------------------------------------', file=f)
                    print('Неучтенка', file=f)
            errors_not = {}
            for element in name_worker:
                try:
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('--------------------------------------', file=f)
                            print(element, file=f)
                    if len(name_worker[element]) != 0:
                        self.status.emit('Формирование файла аналитика руководителя (неучтенка)')
                        if not flag:
                            flag = 1
                            ws.cell(cell_start, 1).value = 'Неучтенные проекты!'
                            ws.cell(cell_start, 1).border = openpyxl.styles.Border(right=thin)
                            ws.cell(cell_start, 1).fill = openpyxl.styles.PatternFill(start_color='f50a0a',
                                                                                      end_color='f50a0a',
                                                                                      fill_type="solid")
                            ws.cell(cell_start, 1).font = openpyxl.styles.Font(name='Calibri', size=14, bold=True,
                                                                               color='ffffff')
                            cell_start += 1
                            row += 1
                        for el in [i for i in name_worker[element]]:
                            if self.log:
                                with open(self.path_worker + '\\log.txt', 'a') as f:
                                    print('--------------------------------------', file=f)
                                    print(el, file=f)
                            len_cell = cell_start
                            while True:
                                try:
                                    if '/' in ws.cell(len_cell, 1).value:
                                        if el == ws.cell(len_cell, 1).value:
                                            len_cell += 1
                                            while True:
                                                if element.partition('_')[2] == ws.cell(len_cell, 1).value:
                                                    row = len_cell
                                                    break
                                                else:
                                                    try:
                                                        if '/' in ws.cell(len_cell, 1).value:
                                                            row = len_cell
                                                            ws.insert_rows(row, amount=1)
                                                            break
                                                        else:
                                                            len_cell += 1
                                                    except TypeError:
                                                        row = len_cell
                                                        break
                                            break
                                        else:
                                            len_cell += 1
                                    else:
                                        len_cell += 1
                                except TypeError:
                                    cell_header = ws.cell(row, 1)
                                    row += 1
                                    cell_header.border = openpyxl.styles.Border(right=thin)
                                    cell_header.value = el
                                    break
                            ws.cell(row, 1).value = element.partition('_')[2]
                            ws.cell(row, 1).border = openpyxl.styles.Border(right=thin)
                            ws.cell(row, 1).alignment = openpyxl.styles.Alignment(horizontal="right")
                            write(ws, row, element, name_worker, el, comments)
                            if element in errors_not:
                                errors_not[element].append(el)
                            else:
                                errors_not[element] = [el]
                            row += 1
                except Exception:
                    errors.append('Ошибка при заполнении файла аналитики в неучтенном проекте ' + element)
                    if self.log:
                        with open(self.path_worker + '\\log.txt', 'a') as f:
                            print('--------------------------------------', file=f)
                            print(traceback.format_exc(), file=f)
                            print('--------------------------------------', file=f)
                            print(traceback.format_exc(), file=f)
            max_row = ws.max_row
            max_col = ws.max_column
            col = 2
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('--------------------------------------', file=f)
                    print('Суммирование', file=f)
            while col <= max_col:
                ws.cell(2, col).value = '= SUM(' + get_column_letter(col) + str(3) + ':' \
                                        + get_column_letter(col) + str(max_row) + ')'
                if ws.cell(1, col).value.weekday() in [5, 6]:
                    for cell_row in range(1, max_row + 1):
                        ws.cell(cell_row, col).fill = openpyxl.styles.PatternFill(start_color='dedede',
                                                                                  end_color='dedede',
                                                                                  fill_type="solid")
                else:
                    ws.cell(2, col).fill = openpyxl.styles.PatternFill(start_color='90e896', end_color='90e896',
                                                                       fill_type="solid")
                col += 1
            self.status.emit('Готово! Открытие файла аналитика руководителя')
            if flag:
                string_ = ''
                for element_ in errors_not:
                    if len(string_) != 0:
                        string_ += '\n'
                    string_ += element_ + ':\n' + '\n'.join(errors_not[element_])
                string_err = 'В файлах работников присутствуют неучтенные проекты!\n' + string_
                self.messageChanged.emit("Скролл", string_err)
            os.chdir(self.path_finish_file)
            if len(errors) != 0:
                self.messageChanged.emit("ВНИМАНИЕ!", '\n'.join(errors))
            try:
                wb.save(filename=name_file)
            except PermissionError:
                self.messageChanged.emit('Вопрос', 'Файл ' + name_file + ' открыт. Для перезаписи файла закройте его '
                                                                         'и нажмите «да». При нажатии «нет» текущая'
                                                                         ' проверка не будет сохранена!')
                while True:
                    val = self.q.get()
                    if val == 1:
                        pass
                    elif val == 2:
                        break
                    elif val == 3:
                        self.status.emit('Прервано пользователем')
                        return
                wb.save(filename=name_file)
            os.startfile(name_file)
        except Exception as e:
            if self.log:
                with open(self.path_worker + '\\log.txt', 'a') as f:
                    print('--------------------------------------', file=f)
                    print(e, file=f)
                    print('--------------------------------------', file=f)
                    print(traceback.format_exc(), file=f)
            else:
                with open(self.path_worker + '\\log.txt', 'w') as f:
                    print(traceback.format_exc(), file=f)
            self.status.emit('Ошибка! Работа программы остановлена')
            return


class CreateFile(QThread):
    status = pyqtSignal(str)
    messageChanged = pyqtSignal(str, str)

    def __init__(self, path_boss, path_worker_list, date, password, year, month, path_finish_file):
        QThread.__init__(self)
        self.path_boss = path_boss
        # self.path_worker = path_worker
        self.path_worker_list = path_worker_list
        self.date = date
        self.password = password
        self.year = year
        self.month = month
        self.path_finish_file = path_finish_file

    def run(self):
        try:
            self.status.emit('Считываем список работников')
            wb = openpyxl.load_workbook(self.path_worker_list)  # Откроем книгу.
            ws = wb.active  # Делаем активным первый лист.
            i = 1
            workers = []
            while True:
                if ws.cell(i, 1).value:
                    workers.append(ws.cell(i, 1).value)
                else:
                    break
                i += 1
            wb.close()

            self.status.emit('Формирование файлов')
            wb = openpyxl.load_workbook(self.path_boss, data_only=True)  # Откроем книгу.
            ws = wb.active  # Делаем активным первый лист.
            name_list = wb.sheetnames[0]
            project_col = None
            j = 1
            while not project_col:
                if ws.cell(1, j).value == 'Шифр объекта':
                    project_col = j
                j += 1
            wb.close()
            try:
                path = self.path_finish_file + '\\' + self.date + 'Табели'
                os.mkdir(path)
            except FileExistsError:
                self.messageChanged.emit('УПС!', 'Папка с именем "' + self.date + '_Табели" уже существует')
                return
            os.chdir(path)
            thin = openpyxl.styles.Side(border_style="thin", color="000000")
            days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            day = days[int(self.date[3]) - 1] if int(self.date[2]) == 0 else days[int(self.date[3:5]) - 1]
            path_book = self.path_boss.replace('\\', '/')
            book_name = path_book.rpartition('/')[2]
            path_book = path_book.rpartition('/')[0]
            for element in workers:
                name_file = self.date + element + '.xlsx'
                self.status.emit('Формирование файла: ' + element)
                wb = openpyxl.Workbook()
                wb.save(filename=name_file)
                wb.close()
                wb = openpyxl.open(filename=name_file, data_only=False, keep_links=True)  # keep_vba=True
                ws = wb.active
                ws.title = 'Рабочий лист'
                wb.create_sheet('Шифры', 2)
                ws = wb['Шифры']
                for i in range(1, 388):
                    link = '=\'' + path_book + '/[' + book_name + ']' + name_list + '\'!$' + \
                           get_column_letter(project_col) + '$' + str(i + 1)
                    # link = '=\'[' + path_book + ']' + name_list + '\'!$' + \
                    #        get_column_letter(project_col) + '$' + str(i + 1)
                    ws.cell(row=i, column=1).value = link
                ws.sheet_state = 'hidden'
                ws.protection.set_password(self.password)
                ws = wb['Рабочий лист']
                ws['A1'] = 'Проект:'
                ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                ws.column_dimensions['A'].width = 40
                ws['A2'].border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                col = 2
                dt = datetime.date(int(self.year), int(self.month), 1)
                for date_for_month in range(1, day + 1):  # Изменяем оформление первой строки.
                    cl = ws.cell(1, col)
                    ws.column_dimensions[get_column_letter(col)].width = 3
                    cl.alignment = openpyxl.styles.Alignment(textRotation=90)
                    cl.border = openpyxl.styles.Border(bottom=thin)
                    # date_day = (dt.replace(day=1) + datetime.timedelta(days=32)).replace(day=date_for_month)
                    # date_day = '{:%Y.%m.%d}'.format(date_day)
                    # try:
                    cl.value = '{:%d.%m.%Y}'.format(
                        (dt.replace(day=1) + datetime.timedelta(days=32)).replace(day=date_for_month))
                    # except ValueError:
                    #     cl.value = '{:%d.%m.%Y}'.format(
                    #         (dt.replace(day=1) + datetime.timedelta(days=32)).replace(day=date_for_month))
                    cl.number_format = 'DD.MM.YYYY'
                    ws.cell(2, col).border = openpyxl.styles.Border(bottom=thin)
                    ws.cell(2, col).value = '= SUM(' + get_column_letter(col) + str(3) + ':' \
                                            + get_column_letter(col) + str(387) + ')'
                    date_ = cl.value.split('.')
                    if datetime.date(int(date_[2]), int(date_[1]), int(date_[0])).weekday() in [5, 6]:
                        for cell_row in range(1, 388):
                            ws.cell(cell_row, col).fill = openpyxl.styles.PatternFill(start_color='dedede',
                                                                                      end_color='dedede',
                                                                                      fill_type="solid")
                    else:
                        ws.cell(2, col).fill = openpyxl.styles.PatternFill(start_color='90e896', end_color='90e896',
                                                                           fill_type="solid")
                    col += 1
                i = 3
                for el in ['_ОТПУСК', '_ПРОСТОЙ', '_Управление (только для руководителя)',
                           '_Отработка листов отклонений', '_ИНОЕ (только по решению руководителя)']:
                    ws.cell(i, 1).value = el
                    ws.cell(i, 1).border = openpyxl.styles.Border(right=thin)
                    i += 1
                for el in range(1, 388):
                    link = '=Шифры!A' + str(el)
                    ws.cell(i, 1).value = link
                    ws.cell(i, 1).border = openpyxl.styles.Border(right=thin)
                    i += 1

                ws.protection.set_password(self.password)
                for row in range(1, ws.max_row + 1):
                    for col in range(2, ws.max_column + 1):
                        ws.cell(row, col).protection = openpyxl.styles.Protection(locked=False)
                wb.save(filename=name_file)
            os.chdir('C:\\')
            self.status.emit('Готово!')
        except Exception as e:
            print(traceback.format_exc())


def about():
    window_add = AboutWindow()
    window_add.exec_()


def default_settings():
    window_add = DefaultWindowWindow()
    window_add.exec_()


class MainWindow(QMainWindow, Main.Ui_MainWindow):  # Главное окно

    ret = pyqtSignal(int)

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        date = [self.dateEdit_start_year, self.dateEdit_finish_month, self.dateEdit_finish_year]
        for element in date:
            element.setDate(datetime.datetime.now())
        date = time.localtime()
        self.dateEdit_start_month.setDate(QDate(date[0], date[1] - 1, 1))
        self.pushButton_path_boss.clicked.connect((lambda: self.browse(1)))
        self.pushButton_path_worker.clicked.connect((lambda: self.browse(2)))
        self.pushButton_path_finish_file.clicked.connect((lambda: self.browse(3)))
        self.pushButton_path_worker_list.clicked.connect((lambda: self.browse(4)))
        self.pushButton_start.clicked.connect(self.start_work)
        self.pushButton_chek_worker.clicked.connect(self.chek_worker)
        self.pushButton_next_month.clicked.connect(self.next_month)
        self.action_default.triggered.connect(default_settings)
        self.action_about.triggered.connect(about)
        with open(pathlib.Path(pathlib.Path.cwd(), 'Настройки.txt'), "r", encoding='utf-8-sig') as f:
            data = json.load(f)
        self.name_eng = ['path_boss', 'path_worker', 'path_finish_file', 'path_worker_list']
        self.line = [self.lineEdit_path_boss, self.lineEdit_path_worker, self.lineEdit_path_finish_file,
                     self.lineEdit_path_worker_list]
        # self.password = None
        for el in data:
            if el != 'password':
                self.line[self.name_eng.index(el)].setText(data[el])
            # else:
            #     self.password = data[el]

    def on_message_changed(self, title, description):
        if title == 'Скролл':
            result = ScrollMessageBox(description)
            result.exec_()
        elif title == 'УПС!':
            QMessageBox.critical(self, title, description)
        elif title == 'ВНИМАНИЕ!':
            QMessageBox.warning(self, title, description)
        elif title == 'Вопрос':
            ans = QMessageBox.question(self, title, description, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                self.thread.q.put(2)
            else:
                self.thread.q.put(3)

    def status(self, value):
        self.statusbar.showMessage(value)

    def stop_thread(self):
        self.thread.quit()
        time.sleep(3)

    def browse(self, num):  # Для кнопки открыть
        if num == 1 or num == 4:
            directory = QFileDialog.getOpenFileName(self, "Find Files", QDir.currentPath())
        else:
            directory = QFileDialog.getExistingDirectory(self, "Find Files", QDir.currentPath())
        if directory:
            if num == 1:
                self.lineEdit_path_boss.setText(directory[0])
            elif num == 2:
                self.lineEdit_path_worker.setText(directory)
            elif num == 3:
                self.lineEdit_path_finish_file.setText(directory)
            elif num == 4:
                self.lineEdit_path_worker_list.setText(directory[0])

    def start_work(self):
        path_boss = self.lineEdit_path_boss.text()
        if not path_boss:
            self.on_message_changed('УПС!', 'Путь к файлу с проектами пуст')
            return
        if os.path.isdir(path_boss):
            self.on_message_changed('УПС!', 'Указанный путь к файлу с проектами является директорией')
            return
        else:
            pass
        path_worker = self.lineEdit_path_worker.text()
        if not path_worker:
            self.on_message_changed('УПС!', 'Путь к файлам сотрудников пуст')
            return
        if os.path.isdir(path_worker):
            pass
        else:
            self.on_message_changed('УПС!', 'Указанный путь к файлам сотрудников не является директорией')
            return
        path_finish_file = self.lineEdit_path_finish_file.text()
        if not path_finish_file:
            self.on_message_changed('УПС!', 'Путь для конечного файла пуст')
            return
        if os.path.isdir(path_finish_file):
            pass
        else:
            self.on_message_changed('УПС!', 'Указанный путь для конечного файла не является директорией')
            return
        all_time = [self.dateEdit_start_year.text(), self.dateEdit_start_month.text(),
                    self.dateEdit_finish_year.text(), self.dateEdit_finish_month.text()]
        if datetime.date(int(all_time[0]), int(all_time[1]), 1) > datetime.date(int(all_time[2]), int(all_time[3]), 1):
            self.on_message_changed('УПС!', 'Время начала проверки больше, чем конечное!')
            return
        path_worker_list = self.lineEdit_path_worker_list.text()
        if not path_worker_list:
            self.on_message_changed('УПС!', 'Путь к списку сотрудников пуст')
            return
        if os.path.isdir(path_worker_list):
            self.on_message_changed('УПС!', 'Указанный путь к списку сотрудников является директорией')
            return
        else:
            pass
        wb = openpyxl.load_workbook(path_worker_list)  # Откроем книгу.
        ws = wb.active  # Делаем активным первый лист.
        i = 1
        workers = []
        while True:
            if ws.cell(i, 1).value:
                workers.append(ws.cell(i, 1).value)
            else:
                break
            i += 1
        wb.close()
        time_text = []
        for el in all_time:
            time_text.append(el)
        date_analytics = [i for i in range(int(time_text[1]), int(time_text[3]) + 1)] \
            if int(time_text[3]) - int(time_text[1]) != 0 else [int(time_text[3])]
        os.chdir(path_worker)
        if len(date_analytics) == 1:
            month_ap = '0' + str(date_analytics[0]) if date_analytics[0] < 10 \
                else str(date_analytics[0])
            current_date = [self.dateEdit_finish_year.text()[2:4] + month_ap]
        else:
            current_date = []
            year = int(self.dateEdit_start_year.text()[2:4])
            month = int(self.dateEdit_start_month.text())
            date_analytics = []
            while True:
                date_analytics.append(month)
                month_ap = '0' + str(month) if month < 10 else str(month)
                current_date.append(str(year) + month_ap)
                if month == int(self.dateEdit_finish_month.text()):
                    if year == int(self.dateEdit_finish_year.text()[2:4]):
                        break
                if month < 12:
                    month += 1
                else:
                    month = 1
                    year += 1
        file_worker = [i for i in os.listdir()]
        # errors = []
        # self.thread = ChekFile(current_date, file_worker, path_worker, workers)
        # self.thread.status.connect(self.status)
        # self.thread.messageChanged.connect(self.on_message_changed)
        # self.thread.start()
        # self.thread.finished.connect(self.stop_thread)
        # for date in current_date:
        #     for folder in file_worker:
        #         err = None
        #         if folder.partition('_')[0] == date:
        #             if (path_worker + "\\" + folder).endswith(".lnk"):
        #                 pass
        #             if os.path.isdir(path_worker + "\\" + folder):
        #                 os.chdir(path_worker + '\\' + folder)
        #                 file_exel = [i.partition('_')[2][:-5] for i in os.listdir() if i.endswith('.xlsm')]
        #                 err = ['Отсутствует табель ' + date + '_' + i for i in workers if i not in file_exel]
        #                 if err:
        #                     errors += err
        #                 for element in [i for i in os.listdir() if i.endswith('.xlsm')]:
        #                     self.status('Проверка заполнения файла ' + element)
        #                     wb = openpyxl.load_workbook(element)
        #                     ws = wb.active
        #                     for el in range(2, ws.max_column):
        #                         try:
        #                             time_year = str(ws.cell(1, el).value.year)[2:]
        #                             time_month = ws.cell(1, el).value.month
        #                             t = time_year + '0' + str(time_month) if time_month < 10 else \
        #                                 time_year + str(time_month)
        #                             if t != date:
        #                                 errors += ['Неверно указано время в файле работника ' + element]
        #                                 break
        #                         except AttributeError:
        #                             break
        # if errors:
        #     self.on_message_changed('УПС!', '\n'.join(errors))
        #     return
        # while self.thread.isRunning():
        #     print('1')
        q = Queue()
        log = 1 if self.action_log.isChecked() else 0
        self.thread = StartProc(path_boss, path_worker, date_analytics, time_text, path_finish_file,
                                self.dateEdit_start_year.text(), self.dateEdit_start_month.text(),
                                self.dateEdit_finish_year.text(), self.dateEdit_finish_month.text(), q, log,
                                path_worker_list)
        self.thread.status.connect(self.status)
        self.thread.messageChanged.connect(self.on_message_changed)
        self.thread.start()
        self.thread.finished.connect(self.stop_thread)

    def chek_worker(self):
        path_worker = self.lineEdit_path_worker.text()
        if not path_worker:
            self.on_message_changed('УПС!', 'Путь к файлам сотрудников пуст')
            return
        if os.path.isdir(path_worker):
            pass
        else:
            self.on_message_changed('УПС!', 'Указанный путь к файлам сотрудников не является директорией')
            return
        path_worker_list = self.lineEdit_path_worker_list.text()
        if not path_worker_list:
            self.on_message_changed('УПС!', 'Путь к списку сотрудников пуст')
            return
        if os.path.isdir(path_worker_list):
            self.on_message_changed('УПС!', 'Указанный путь к списку сотрудников является директорией')
            return
        else:
            pass
        wb = openpyxl.load_workbook(path_worker_list)  # Откроем книгу.
        ws = wb.active  # Делаем активным первый лист.
        i = 1
        workers = []
        while True:
            if ws.cell(i, 1).value:
                workers.append(ws.cell(i, 1).value)
            else:
                break
            i += 1
        wb.close()
        path_finish_file = self.lineEdit_path_finish_file.text()
        if not path_finish_file:
            self.on_message_changed('УПС!', 'Путь для конечного файла пуст')
            return
        if os.path.isdir(path_finish_file):
            pass
        else:
            self.on_message_changed('УПС!', 'Указанный путь для конечного файла не является директорией')
            return
        time_now = datetime.date.today()
        if time_now.strftime('%d') == 1:
            if int(time_now.strftime('%m')) - 1 == 0:
                start_date = datetime.date(int(time_now.strftime('%Y')), 12, 1)
            else:
                start_date = datetime.date(int(time_now.strftime('%Y')), int(time_now.strftime('%m')) - 1, 1)
        else:
            start_date = datetime.date(int(time_now.strftime('%Y')), int(time_now.strftime('%m')), 1)
        finish_date = datetime.date(int(time_now.strftime('%Y')), int(time_now.strftime('%m')),
                                    int(time_now.strftime('%d')))
        date_analytics = start_date.strftime('%y') + start_date.strftime('%m')
        self.thread = WorkerAnalytics(path_worker, path_worker_list, date_analytics, path_finish_file, finish_date)
        self.thread.status.connect(self.status)
        self.thread.messageChanged.connect(self.on_message_changed)
        self.thread.start()
        self.thread.finished.connect(self.stop_thread)

    def next_month(self):
        path_boss = self.lineEdit_path_boss.text()
        if not path_boss:
            self.on_message_changed('УПС!', 'Путь к файлу с проектами пуст')
            return
        if os.path.isdir(path_boss):
            self.on_message_changed('УПС!', 'Указанный путь к файлу с проектами является директорией')
            return
        else:
            pass
        # path_worker = self.lineEdit_path_worker.text()
        # if not path_worker:
        #     self.on_message_changed('УПС!', 'Путь к файлам сотрудников пуст')
        #     return
        # if os.path.isdir(path_worker):
        #     pass
        # else:
        #     self.on_message_changed('УПС!', 'Указанный путь к файлам сотрудников не является директорией')
        #     return
        path_worker_list = self.lineEdit_path_worker_list.text()
        if not path_worker_list:
            self.on_message_changed('УПС!', 'Путь к списку сотрудников пуст')
            return
        if os.path.isdir(path_worker_list):
            self.on_message_changed('УПС!', 'Указанный путь к списку сотрудников является директорией')
            return
        else:
            pass
        # if not self.password:
        with open(pathlib.Path(pathlib.Path.cwd(), 'Настройки.txt'), "r", encoding='utf-8-sig') as f:
            data = json.load(f)
        if 'password' in data:
            password = data['password']
        else:
            self.on_message_changed('УПС!', 'Отсутствует пароль')
            return
        # else:
        #     password = self.password
        path_finish_file = self.lineEdit_path_finish_file.text()
        if not path_finish_file:
            self.on_message_changed('УПС!', 'Путь для конечного файла пуст')
            return
        if os.path.isdir(path_finish_file):
            pass
        else:
            self.on_message_changed('УПС!', 'Указанный путь для конечного файла не является директорией')
            return
        year, month = self.dateEdit_finish_year.text(), self.dateEdit_finish_month.text()
        if int(month) == 12:
            year = str(int(year[2:]) + 1)
            month = '01'
        elif int(month) < 10:
            year = year[2:]
            month = int(month[1]) + 1
            if month != 10:
                month = '0' + str(month)
        else:
            year = year[2:]
            month = str(int(month) + 1)
        date = year + month + '_'
        self.thread = CreateFile(path_boss, path_worker_list, date, password,
                                 self.dateEdit_finish_year.text(), self.dateEdit_finish_month.text(), path_finish_file)
        self.thread.status.connect(self.status)
        self.thread.messageChanged.connect(self.on_message_changed)
        self.thread.start()
        self.thread.finished.connect(self.stop_thread)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
